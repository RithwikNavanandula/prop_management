"""Property CRUD routes."""
import logging
import csv
import io
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from typing import Optional
from decimal import Decimal, InvalidOperation
from app.database import get_db
from app.auth.dependencies import get_current_user, require_permissions
from app.auth.models import UserAccount
from app.modules.properties.models import (
    Property, Building, Floor, Unit, Asset, UnitAsset, Owner, Tenant, Vendor,
    PropertyOwnerLink, Region, TenantOrg, StaffUser
)
from app.utils.qrcode_service import generate_qr_code
from app.modules.compliance.models import Document
import os
import shutil
from datetime import datetime, date

logger = logging.getLogger(__name__)
router = APIRouter(
    prefix="/api/properties",
    tags=["Properties"],
    dependencies=[Depends(require_permissions(["properties", "portfolio"]))],
)

NULL_LIKE_STRINGS = {"", "null", "none", "nan"}


def _coerce_column_value(column, value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        if value.lower() in NULL_LIKE_STRINGS:
            return None

    try:
        python_type = column.type.python_type
    except (NotImplementedError, AttributeError):
        return value

    if python_type is bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        if isinstance(value, str):
            lowered = value.lower()
            if lowered in {"true", "1", "yes", "y", "on"}:
                return True
            if lowered in {"false", "0", "no", "n", "off"}:
                return False
        raise ValueError("invalid boolean")

    if python_type in (int, float, Decimal):
        try:
            return python_type(value)
        except (TypeError, ValueError, InvalidOperation) as exc:
            raise ValueError("invalid numeric value") from exc

    if python_type is datetime:
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            except ValueError as exc:
                raise ValueError("invalid datetime") from exc
        raise ValueError("invalid datetime")

    if python_type is date:
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value).date()
            except ValueError:
                try:
                    return datetime.strptime(value, "%Y-%m-%d").date()
                except ValueError as exc:
                    raise ValueError("invalid date") from exc
        raise ValueError("invalid date")

    return value


def _sanitize_model_payload(model, data: dict, blocked_fields: set | None = None) -> dict:
    blocked_fields = blocked_fields or set()
    cleaned = {}
    errors = []

    for key, value in data.items():
        if key in blocked_fields:
            continue
        column = model.__table__.columns.get(key)
        if column is None:
            continue
        try:
            cleaned[key] = _coerce_column_value(column, value)
        except ValueError:
            errors.append(key)

    if errors:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid values for fields: {', '.join(errors)}"
        )
    return cleaned


@router.get("/tenant-orgs")
def list_tenant_orgs(db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    items = db.query(TenantOrg).all()
    return {"total": len(items), "items": [_org_dict(o) for o in items]}


@router.get("")
def list_properties(
    search: Optional[str] = None,
    property_type: Optional[str] = None,
    status: Optional[str] = "Active",
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    q = db.query(Property).filter(Property.is_deleted == False)
    # Multi-tenant isolation
    if user.tenant_org_id:
        q = q.filter(Property.tenant_org_id == user.tenant_org_id)
    if status:
        q = q.filter(Property.status == status)
    if property_type:
        q = q.filter(Property.property_type == property_type)
    if search:
        q = q.filter(or_(
            Property.property_name.ilike(f"%{search}%"),
            Property.property_code.ilike(f"%{search}%"),
            Property.city.ilike(f"%{search}%"),
        ))
    total = q.count()
    items = q.order_by(Property.id.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_prop_dict(p) for p in items]}


@router.post("", status_code=201)
def create_property(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    clean_data = _sanitize_model_payload(Property, data, blocked_fields={"id", "created_at", "updated_at", "created_by", "updated_by"})

    # Enforce tenant scope automatically for tenant-scoped users.
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id

    prop = Property(**clean_data)
    prop.created_by = user.id
    db.add(prop)
    db.commit()
    db.refresh(prop)
    return _prop_dict(prop)


@router.get("/{prop_id}")
def get_property(prop_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    prop = db.query(Property).filter(Property.id == prop_id, Property.is_deleted == False).first()
    if not prop:
        raise HTTPException(404, "Property not found")
    return _prop_dict(prop)


@router.put("/{prop_id}")
def update_property(prop_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    prop = db.query(Property).filter(Property.id == prop_id).first()
    if not prop:
        raise HTTPException(404, "Property not found")

    # Prevent cross-tenant mutation for tenant-scoped users.
    if user.tenant_org_id and prop.tenant_org_id != user.tenant_org_id:
        raise HTTPException(404, "Property not found")

    clean_data = _sanitize_model_payload(Property, data, blocked_fields={"id", "created_at", "updated_at", "created_by", "updated_by"})
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id

    for k, v in clean_data.items():
        setattr(prop, k, v)

    prop.updated_by = user.id
    db.commit()
    db.refresh(prop)
    return _prop_dict(prop)


@router.delete("/{prop_id}")
def delete_property(prop_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    prop = db.query(Property).filter(Property.id == prop_id).first()
    if not prop:
        raise HTTPException(404, "Property not found")
    prop.is_deleted = True
    prop.status = "Inactive"
    db.commit()
    return {"message": "Property deleted"}


# --- Units ---
@router.get("/{prop_id}/units")
def list_units(prop_id: int, status: Optional[str] = None, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Unit).filter(Unit.property_id == prop_id, Unit.is_deleted == False)
    if status:
        q = q.filter(Unit.current_status == status)
    items = q.order_by(Unit.unit_number).all()
    return {"total": len(items), "items": [_unit_dict(u) for u in items]}


@router.get("/{prop_id}/units/{unit_id}")
def get_unit(prop_id: int, unit_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    unit = db.query(Unit).filter(Unit.id == unit_id, Unit.property_id == prop_id, Unit.is_deleted == False).first()
    if not unit:
        raise HTTPException(404, "Unit not found")
    return _unit_dict(unit)


@router.post("/{prop_id}/units", status_code=201)
def create_unit(prop_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    unit_data = _sanitize_model_payload(
        Unit,
        data,
        blocked_fields={"id", "created_at", "updated_at", "property_id", "created_by", "updated_by"},
    )
    if not unit_data.get("unit_number"):
        raise HTTPException(status_code=422, detail="unit_number is required")
    if user.tenant_org_id:
        unit_data["tenant_org_id"] = user.tenant_org_id
    unit = Unit(**unit_data)
    unit.property_id = prop_id
    unit.created_by = user.id
    db.add(unit)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=422, detail=f"Invalid unit payload: {exc.orig}") from exc
    db.refresh(unit)
    return _unit_dict(unit)


@router.put("/{prop_id}/units/{unit_id}")
def update_unit(prop_id: int, unit_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    unit = db.query(Unit).filter(Unit.id == unit_id, Unit.property_id == prop_id).first()
    if not unit:
        raise HTTPException(404, "Unit not found")
    unit_data = _sanitize_model_payload(
        Unit,
        data,
        blocked_fields={"id", "created_at", "updated_at", "property_id", "created_by", "updated_by"},
    )
    if user.tenant_org_id:
        unit_data["tenant_org_id"] = user.tenant_org_id
    for k, v in unit_data.items():
        setattr(unit, k, v)
    unit.updated_by = user.id
    db.commit()
    db.refresh(unit)
    return _unit_dict(unit)


@router.delete("/{prop_id}/units/{unit_id}")
def delete_unit(prop_id: int, unit_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    unit = db.query(Unit).filter(Unit.id == unit_id, Unit.property_id == prop_id).first()
    if not unit:
        raise HTTPException(404, "Unit not found")
    unit.is_deleted = True
    unit.status = "Inactive"
    db.commit()
    return {"message": "Unit deleted"}


@router.post("/{prop_id}/units/{unit_id}/qrcode")
def generate_unit_qrcode(prop_id: int, unit_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    unit = db.query(Unit).filter(Unit.id == unit_id, Unit.property_id == prop_id).first()
    if not unit:
        raise HTTPException(404, "Unit not found")
        
    qr_content = f"UNIT:{unit.unit_number}"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"unit_{unit.id}_{timestamp}.png"
    
    try:
        url = generate_qr_code(qr_content, filename)
        unit.qr_code_image_url = url
        unit.qr_code_value = qr_content
        unit.qr_code_last_generated_at = datetime.now()
        db.commit()
        db.refresh(unit)
        return _unit_dict(unit)
    except Exception as e:
        raise HTTPException(500, f"Failed to generate QR code: {str(e)}")


# --- Buildings ---
@router.get("/{prop_id}/buildings")
def list_buildings(prop_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    items = db.query(Building).filter(Building.property_id == prop_id, Building.is_deleted == False).all()
    return {"total": len(items), "items": [_bldg_dict(b) for b in items]}


@router.post("/{prop_id}/buildings", status_code=201)
def create_building(prop_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    bldg = Building(**{k: v for k, v in data.items() if hasattr(Building, k)})
    bldg.property_id = prop_id
    bldg.created_by = user.id
    db.add(bldg)
    db.commit()
    db.refresh(bldg)
    return _bldg_dict(bldg)


@router.put("/{prop_id}/buildings/{bldg_id}")
def update_building(prop_id: int, bldg_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    bldg = db.query(Building).filter(Building.id == bldg_id, Building.property_id == prop_id).first()
    if not bldg:
        raise HTTPException(404, "Building not found")
    for k, v in data.items():
        if hasattr(bldg, k) and k not in ("id", "created_at", "property_id"):
            setattr(bldg, k, v)
    bldg.updated_by = user.id
    db.commit()
    db.refresh(bldg)
    return _bldg_dict(bldg)


@router.delete("/{prop_id}/buildings/{bldg_id}")
def delete_building(prop_id: int, bldg_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    bldg = db.query(Building).filter(Building.id == bldg_id, Building.property_id == prop_id).first()
    if not bldg:
        raise HTTPException(404, "Building not found")
    bldg.is_deleted = True
    bldg.status = "Inactive"
    db.commit()
    return {"message": "Building deleted"}


# --- Floors ---
@router.get("/{prop_id}/buildings/{bldg_id}/floors")
def list_floors(prop_id: int, bldg_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    items = db.query(Floor).filter(Floor.building_id == bldg_id).all()
    return {"total": len(items), "items": [_floor_dict(f) for f in items]}


@router.post("/{prop_id}/buildings/{bldg_id}/floors", status_code=201)
def create_floor(prop_id: int, bldg_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    floor = Floor(**{k: v for k, v in data.items() if hasattr(Floor, k)})
    floor.building_id = bldg_id
    db.add(floor)
    db.commit()
    db.refresh(floor)
    return _floor_dict(floor)


@router.put("/{prop_id}/buildings/{bldg_id}/floors/{floor_id}")
def update_floor(prop_id: int, bldg_id: int, floor_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    floor = db.query(Floor).filter(Floor.id == floor_id, Floor.building_id == bldg_id).first()
    if not floor:
        raise HTTPException(404, "Floor not found")
    for k, v in data.items():
        if hasattr(floor, k) and k not in ("id", "created_at", "building_id"):
            setattr(floor, k, v)
    db.commit()
    db.refresh(floor)
    return _floor_dict(floor)


@router.delete("/{prop_id}/buildings/{bldg_id}/floors/{floor_id}")
def delete_floor(prop_id: int, bldg_id: int, floor_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    floor = db.query(Floor).filter(Floor.id == floor_id, Floor.building_id == bldg_id).first()
    if not floor:
        raise HTTPException(404, "Floor not found")
    db.delete(floor)
    db.commit()
    return {"message": "Floor deleted"}


def _normalized_row_keys(row: dict) -> dict:
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized_key = str(key).strip().lower().replace(" ", "_").replace("-", "_")
        normalized[normalized_key] = value
    return normalized


def _row_value(row: dict, *keys):
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                continue
        return value
    return None


def _iter_sheet_rows(file_bytes: bytes, filename: str) -> tuple[list[dict], list[dict]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        decoded = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        units_rows = [_normalized_row_keys(dict(row)) for row in reader]
        return [], units_rows

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise HTTPException(
            status_code=500,
            detail="Excel import requires openpyxl. Install dependencies from requirements.txt",
        ) from exc

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

    def sheet_to_rows(sheet_name: str) -> list[dict]:
        if sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[sheet_name]
        values = list(ws.values)
        if not values:
            return []
        header = [str(v).strip() if v is not None else "" for v in values[0]]
        rows = []
        for raw in values[1:]:
            row = {}
            for idx, column_name in enumerate(header):
                if not column_name:
                    continue
                row[column_name] = raw[idx] if idx < len(raw) else None
            rows.append(_normalized_row_keys(row))
        return rows

    building_rows = sheet_to_rows("Buildings")
    units_rows = sheet_to_rows("Units")
    if not units_rows and workbook.sheetnames:
        units_rows = sheet_to_rows(workbook.sheetnames[0])
    return building_rows, units_rows


@router.post("/{prop_id}/import/buildings-units")
async def import_buildings_and_units(
    prop_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    prop = db.query(Property).filter(Property.id == prop_id, Property.is_deleted == False).first()
    if not prop:
        raise HTTPException(404, "Property not found")

    if user.tenant_org_id and prop.tenant_org_id != user.tenant_org_id:
        raise HTTPException(404, "Property not found")

    if not file.filename:
        raise HTTPException(400, "File name is required")
    lower_name = file.filename.lower()
    if not (lower_name.endswith(".csv") or lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm")):
        raise HTTPException(400, "Unsupported file type. Use .csv or .xlsx")

    file_bytes = await file.read()
    building_rows, unit_rows = _iter_sheet_rows(file_bytes, file.filename)
    if not unit_rows and not building_rows:
        raise HTTPException(400, "No rows found in import file")

    tenant_org_id = prop.tenant_org_id or user.tenant_org_id
    building_cache: dict[str, Building] = {}
    floor_cache: dict[tuple[int, int], Floor] = {}
    errors = []
    stats = {
        "buildings_created": 0,
        "buildings_updated": 0,
        "floors_created": 0,
        "units_created": 0,
        "units_updated": 0,
    }

    def upsert_building(row: dict, row_index: int) -> Optional[Building]:
        code = _row_value(row, "building_code", "code")
        name = _row_value(row, "building_name", "name")
        if not code and not name:
            return None

        cache_key = (str(code).strip().upper() if code else str(name).strip().lower())
        if cache_key in building_cache:
            return building_cache[cache_key]

        q = db.query(Building).filter(Building.property_id == prop_id, Building.is_deleted == False)
        if code:
            q = q.filter(Building.building_code == str(code).strip())
        else:
            q = q.filter(Building.building_name == str(name).strip())
        building = q.first()

        payload = {
            "building_code": str(code).strip() if code else str(name).strip().upper().replace(" ", "-"),
            "building_name": str(name).strip() if name else str(code).strip(),
            "floors_count": _row_value(row, "floors_count"),
            "year_built": _row_value(row, "year_built"),
            "status": _row_value(row, "status", "building_status"),
            "tenant_org_id": tenant_org_id,
        }
        clean_payload = _sanitize_model_payload(
            Building,
            payload,
            blocked_fields={"id", "created_at", "updated_at", "property_id", "created_by", "updated_by"},
        )

        if building:
            for key, value in clean_payload.items():
                setattr(building, key, value)
            building.updated_by = user.id
            stats["buildings_updated"] += 1
        else:
            building = Building(**clean_payload)
            building.property_id = prop_id
            building.created_by = user.id
            db.add(building)
            db.flush()
            stats["buildings_created"] += 1

        building_cache[cache_key] = building
        return building

    for idx, row in enumerate(building_rows, start=2):
        try:
            upsert_building(row, idx)
        except Exception as exc:
            errors.append({"row": idx, "section": "buildings", "error": str(exc)})

    for idx, row in enumerate(unit_rows, start=2):
        try:
            unit_number = _row_value(row, "unit_number", "unit_no")
            if not unit_number:
                continue

            building = upsert_building(row, idx)
            floor_id = None
            if building:
                floor_number = _row_value(row, "floor_number")
                if floor_number is not None:
                    floor_number_int = int(floor_number)
                    floor_key = (building.id, floor_number_int)
                    floor = floor_cache.get(floor_key)
                    if not floor:
                        floor = db.query(Floor).filter(
                            Floor.building_id == building.id,
                            Floor.floor_number == floor_number_int,
                        ).first()
                    if not floor:
                        floor_payload = _sanitize_model_payload(
                            Floor,
                            {
                                "tenant_org_id": tenant_org_id,
                                "floor_number": floor_number_int,
                                "floor_name": _row_value(row, "floor_name"),
                                "status": _row_value(row, "floor_status", "status"),
                            },
                            blocked_fields={"id", "created_at", "updated_at", "building_id"},
                        )
                        floor = Floor(**floor_payload)
                        floor.building_id = building.id
                        db.add(floor)
                        db.flush()
                        stats["floors_created"] += 1
                    floor_cache[floor_key] = floor
                    floor_id = floor.id

            unit_payload = {
                "tenant_org_id": tenant_org_id,
                "building_id": building.id if building else None,
                "floor_id": floor_id,
                "unit_number": str(unit_number).strip(),
                "unit_name": _row_value(row, "unit_name"),
                "unit_type": _row_value(row, "unit_type"),
                "area_sqft": _row_value(row, "area_sqft"),
                "area_sqm": _row_value(row, "area_sqm"),
                "bedrooms": _row_value(row, "bedrooms"),
                "bathrooms": _row_value(row, "bathrooms"),
                "rooms": _row_value(row, "rooms"),
                "balconies": _row_value(row, "balconies"),
                "parking_slots": _row_value(row, "parking_slots"),
                "ceiling_height_ft": _row_value(row, "ceiling_height_ft"),
                "load_capacity_tons": _row_value(row, "load_capacity_tons"),
                "hvac_type": _row_value(row, "hvac_type"),
                "current_status": _row_value(row, "current_status"),
                "market_rent": _row_value(row, "market_rent"),
                "min_lease_term": _row_value(row, "min_lease_term"),
                "max_occupancy": _row_value(row, "max_occupancy"),
                "usage_type": _row_value(row, "usage_type"),
                "status": _row_value(row, "status"),
                "photo_url": _row_value(row, "photo_url"),
                "description": _row_value(row, "description"),
            }
            clean_unit_payload = _sanitize_model_payload(
                Unit,
                unit_payload,
                blocked_fields={"id", "created_at", "updated_at", "property_id", "created_by", "updated_by"},
            )

            uq = db.query(Unit).filter(Unit.property_id == prop_id, Unit.unit_number == clean_unit_payload["unit_number"])
            if clean_unit_payload.get("building_id"):
                uq = uq.filter(Unit.building_id == clean_unit_payload["building_id"])
            unit = uq.first()
            if unit:
                for key, value in clean_unit_payload.items():
                    setattr(unit, key, value)
                unit.updated_by = user.id
                stats["units_updated"] += 1
            else:
                unit = Unit(**clean_unit_payload)
                unit.property_id = prop_id
                unit.created_by = user.id
                db.add(unit)
                stats["units_created"] += 1
        except Exception as exc:
            errors.append({"row": idx, "section": "units", "error": str(exc)})

    db.commit()
    return {
        "message": "Import completed",
        "imported_rows": len(unit_rows),
        "building_rows": len(building_rows),
        "stats": stats,
        "errors": errors[:100],
        "error_count": len(errors),
    }


# --- Unit Assets (nested under properties – backward compat) ---
@router.get("/{prop_id}/units/{unit_id}/assets")
def list_unit_assets(prop_id: int, unit_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    items = db.query(Asset).filter(Asset.unit_id == unit_id).all()
    return {"total": len(items), "items": [_asset_dict(a) for a in items]}


@router.post("/{prop_id}/units/{unit_id}/assets", status_code=201)
def create_unit_asset(prop_id: int, unit_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    asset = Asset(**{k: v for k, v in data.items() if hasattr(Asset, k)})
    asset.unit_id = unit_id
    asset.property_id = prop_id
    asset.allocated_at = datetime.now()
    if not asset.asset_number:
        count = db.query(Asset).count()
        asset.asset_number = f"AST-{count + 1:05d}"
    if user.tenant_org_id:
        asset.tenant_org_id = user.tenant_org_id
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return _asset_dict(asset)


# --- Documents ---
@router.get("/{prop_id}/units/{unit_id}/documents")
def list_unit_documents(prop_id: int, unit_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    docs = db.query(Document).filter(Document.owner_entity_type == "Unit", Document.owner_entity_id == unit_id).all()
    return {"total": len(docs), "items": [_doc_dict(x) for x in docs]}


@router.post("/{prop_id}/units/{unit_id}/documents", status_code=201)
async def upload_unit_document(prop_id: int, unit_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    from app.config import get_settings
    settings = get_settings()
    
    upload_path = os.path.join(settings.UPLOAD_DIR, f"prop_{prop_id}", f"unit_{unit_id}")
    os.makedirs(upload_path, exist_ok=True)
    
    file_path = os.path.join(upload_path, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    doc = Document(
        tenant_org_id=user.tenant_org_id,
        owner_entity_type="Unit",
        owner_entity_id=unit_id,
        file_name=file.filename,
        file_path=file_path.replace("\\", "/"),
        mime_type=file.content_type,
        upload_date=datetime.now()
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return _doc_dict(doc)


def _doc_dict(d):
    return {c.name: getattr(d, c.name) for c in d.__table__.columns}


# --- Tenants ---
tenants_router = APIRouter(
    prefix="/api/tenants",
    tags=["Tenants"],
    dependencies=[Depends(require_permissions(["tenants", "leases", "portfolio"]))],
)


@tenants_router.get("")
def list_tenants(search: Optional[str] = None, skip: int = 0, limit: int = 50,
                 db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Tenant).filter(Tenant.is_deleted == False)
    # Multi-tenant isolation
    if user.tenant_org_id:
        q = q.filter(Tenant.tenant_org_id == user.tenant_org_id)
    if search:
        q = q.filter(or_(Tenant.first_name.ilike(f"%{search}%"), Tenant.last_name.ilike(f"%{search}%"),
                         Tenant.email.ilike(f"%{search}%"), Tenant.tenant_code.ilike(f"%{search}%")))
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    return {"total": total, "items": [_tenant_dict(t) for t in items]}


@tenants_router.post("", status_code=201)
def create_tenant(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    clean_data = _sanitize_model_payload(
        Tenant,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    tenant = Tenant(**clean_data)
    db.add(tenant)
    db.commit()
    db.refresh(tenant)
    return _tenant_dict(tenant)


@tenants_router.get("/{tenant_id}")
def get_tenant(tenant_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Tenant).filter(Tenant.id == tenant_id, Tenant.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Tenant.tenant_org_id == user.tenant_org_id)
    t = q.first()
    if not t:
        raise HTTPException(404, "Tenant not found")
    return _tenant_dict(t)


@tenants_router.put("/{tenant_id}")
def update_tenant(tenant_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Tenant).filter(Tenant.id == tenant_id, Tenant.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Tenant.tenant_org_id == user.tenant_org_id)
    t = q.first()
    if not t:
        raise HTTPException(404, "Tenant not found")
    clean_data = _sanitize_model_payload(
        Tenant,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    for k, v in clean_data.items():
        setattr(t, k, v)
    db.commit()
    db.refresh(t)
    return _tenant_dict(t)


@tenants_router.delete("/{tenant_id}")
def delete_tenant(tenant_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Tenant).filter(Tenant.id == tenant_id, Tenant.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Tenant.tenant_org_id == user.tenant_org_id)
    t = q.first()
    if not t:
        raise HTTPException(404, "Tenant not found")
    t.is_deleted = True
    db.commit()
    return {"message": "Tenant deleted"}


# --- Staff ---
staff_router = APIRouter(
    prefix="/api/staff",
    tags=["Staff"],
    dependencies=[Depends(require_permissions(["admin"]))],
)


@staff_router.get("")
def list_staff(
    search: Optional[str] = None,
    role_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    q = db.query(StaffUser)
    if user.tenant_org_id:
        q = q.filter(StaffUser.tenant_org_id == user.tenant_org_id)
    if search:
        q = q.filter(or_(
            StaffUser.employee_code.ilike(f"%{search}%"),
            StaffUser.first_name.ilike(f"%{search}%"),
            StaffUser.last_name.ilike(f"%{search}%"),
            StaffUser.email.ilike(f"%{search}%"),
            StaffUser.department.ilike(f"%{search}%"),
        ))
    if role_id:
        q = q.filter(StaffUser.role_id == role_id)
    if status:
        q = q.filter(StaffUser.status == status)
    total = q.count()
    items = q.order_by(StaffUser.id.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_staff_dict(s) for s in items]}


@staff_router.post("", status_code=201)
def create_staff(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    clean_data = _sanitize_model_payload(
        StaffUser,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    if not clean_data.get("employee_code") or not clean_data.get("first_name"):
        raise HTTPException(status_code=422, detail="employee_code and first_name are required")
    duplicate = db.query(StaffUser).filter(
        StaffUser.tenant_org_id == clean_data.get("tenant_org_id"),
        StaffUser.employee_code == clean_data["employee_code"],
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="employee_code already exists")
    staff = StaffUser(**clean_data)
    db.add(staff)
    db.commit()
    db.refresh(staff)
    return _staff_dict(staff)


@staff_router.get("/{staff_id}")
def get_staff(staff_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(StaffUser).filter(StaffUser.id == staff_id)
    if user.tenant_org_id:
        q = q.filter(StaffUser.tenant_org_id == user.tenant_org_id)
    staff = q.first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    return _staff_dict(staff)


@staff_router.put("/{staff_id}")
def update_staff(staff_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(StaffUser).filter(StaffUser.id == staff_id)
    if user.tenant_org_id:
        q = q.filter(StaffUser.tenant_org_id == user.tenant_org_id)
    staff = q.first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")

    clean_data = _sanitize_model_payload(
        StaffUser,
        data,
        blocked_fields={"id", "created_at", "updated_at", "tenant_org_id"},
    )
    if "employee_code" in clean_data:
        duplicate = db.query(StaffUser).filter(
            StaffUser.id != staff.id,
            StaffUser.tenant_org_id == staff.tenant_org_id,
            StaffUser.employee_code == clean_data["employee_code"],
        ).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="employee_code already exists")

    for k, v in clean_data.items():
        setattr(staff, k, v)

    # Keep linked user role in sync if staff role is changed.
    if "role_id" in clean_data:
        uq = db.query(UserAccount).filter(
            UserAccount.linked_entity_type == "Staff",
            UserAccount.linked_entity_id == staff.id,
        )
        if user.tenant_org_id:
            uq = uq.filter(UserAccount.tenant_org_id == user.tenant_org_id)
        for linked_user in uq.all():
            linked_user.role_id = clean_data["role_id"]

    db.commit()
    db.refresh(staff)
    return _staff_dict(staff)


@staff_router.delete("/{staff_id}")
def delete_staff(staff_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(StaffUser).filter(StaffUser.id == staff_id)
    if user.tenant_org_id:
        q = q.filter(StaffUser.tenant_org_id == user.tenant_org_id)
    staff = q.first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")

    # Prefer soft delete to avoid breaking maintenance/work-order references.
    staff.status = "Inactive"
    uq = db.query(UserAccount).filter(
        UserAccount.linked_entity_type == "Staff",
        UserAccount.linked_entity_id == staff.id,
    )
    if user.tenant_org_id:
        uq = uq.filter(UserAccount.tenant_org_id == user.tenant_org_id)
    for linked_user in uq.all():
        linked_user.is_active = False

    db.commit()
    return {"message": "Staff deactivated"}


# --- Owners ---
owners_router = APIRouter(
    prefix="/api/owners",
    tags=["Owners"],
    dependencies=[Depends(require_permissions(["owners", "portfolio"]))],
)


@owners_router.get("")
def list_owners(db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Owner).filter(Owner.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Owner.tenant_org_id == user.tenant_org_id)
    items = q.all()
    return {"total": len(items), "items": [_owner_dict(o) for o in items]}


@owners_router.post("", status_code=201)
def create_owner(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    clean_data = _sanitize_model_payload(
        Owner,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    owner = Owner(**clean_data)
    db.add(owner)
    db.commit()
    db.refresh(owner)
    return _owner_dict(owner)


@owners_router.get("/{owner_id}")
def get_owner(owner_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Owner).filter(Owner.id == owner_id, Owner.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Owner.tenant_org_id == user.tenant_org_id)
    o = q.first()
    if not o:
        raise HTTPException(404, "Owner not found")
    return _owner_dict(o)


@owners_router.put("/{owner_id}")
def update_owner(owner_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Owner).filter(Owner.id == owner_id, Owner.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Owner.tenant_org_id == user.tenant_org_id)
    o = q.first()
    if not o:
        raise HTTPException(404, "Owner not found")
    clean_data = _sanitize_model_payload(
        Owner,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    for k, v in clean_data.items():
        setattr(o, k, v)
    db.commit()
    db.refresh(o)
    return _owner_dict(o)


@owners_router.delete("/{owner_id}")
def delete_owner(owner_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Owner).filter(Owner.id == owner_id, Owner.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Owner.tenant_org_id == user.tenant_org_id)
    o = q.first()
    if not o:
        raise HTTPException(404, "Owner not found")
    o.is_deleted = True
    db.commit()
    return {"message": "Owner deleted"}


# --- Vendors ---
vendors_router = APIRouter(
    prefix="/api/vendors",
    tags=["Vendors"],
    dependencies=[Depends(require_permissions(["vendors", "maintenance", "work_orders"]))],
)


@vendors_router.get("")
def list_vendors(db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Vendor).filter(Vendor.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Vendor.tenant_org_id == user.tenant_org_id)
    items = q.all()
    return {"total": len(items), "items": [_v_dict(v) for v in items]}


@vendors_router.post("", status_code=201)
def create_vendor(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    clean_data = _sanitize_model_payload(
        Vendor,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    vendor = Vendor(**clean_data)
    db.add(vendor)
    db.commit()
    db.refresh(vendor)
    return _v_dict(vendor)


@vendors_router.get("/{vendor_id}")
def get_vendor(vendor_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Vendor).filter(Vendor.id == vendor_id, Vendor.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Vendor.tenant_org_id == user.tenant_org_id)
    v = q.first()
    if not v:
        raise HTTPException(404, "Vendor not found")
    return _v_dict(v)


@vendors_router.put("/{vendor_id}")
def update_vendor(vendor_id: int, data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Vendor).filter(Vendor.id == vendor_id, Vendor.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Vendor.tenant_org_id == user.tenant_org_id)
    v = q.first()
    if not v:
        raise HTTPException(404, "Vendor not found")
    clean_data = _sanitize_model_payload(
        Vendor,
        data,
        blocked_fields={"id", "created_at", "updated_at"},
    )
    if user.tenant_org_id:
        clean_data["tenant_org_id"] = user.tenant_org_id
    for k, v_val in clean_data.items():
        setattr(v, k, v_val)
    db.commit()
    db.refresh(v)
    return _v_dict(v)


@vendors_router.delete("/{vendor_id}")
def delete_vendor(vendor_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    q = db.query(Vendor).filter(Vendor.id == vendor_id, Vendor.is_deleted == False)
    if user.tenant_org_id:
        q = q.filter(Vendor.tenant_org_id == user.tenant_org_id)
    v = q.first()
    if not v:
        raise HTTPException(404, "Vendor not found")
    v.is_deleted = True
    db.commit()
    return {"message": "Vendor deleted"}


# --- Regions ---
@router.get("/regions")
def list_regions(db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    items = db.query(Region).all()
    return {"total": len(items), "items": [{c.name: getattr(r, c.name) for c in r.__table__.columns} for r in items]}


@router.post("/regions", status_code=201)
def create_region(data: dict, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    region = Region(**{k: v for k, v in data.items() if hasattr(Region, k)})
    db.add(region)
    db.commit()
    db.refresh(region)
    return {c.name: getattr(region, c.name) for c in region.__table__.columns}


# --- Helpers ---
def _prop_dict(p):
    return {c.name: getattr(p, c.name) for c in p.__table__.columns}

def _unit_dict(u):
    return {c.name: getattr(u, c.name) for c in u.__table__.columns}

def _bldg_dict(b):
    return {c.name: getattr(b, c.name) for c in b.__table__.columns}

def _floor_dict(f):
    return {c.name: getattr(f, c.name) for c in f.__table__.columns}

def _asset_dict(a):
    return {c.name: getattr(a, c.name) for c in a.__table__.columns}

def _tenant_dict(t):
    return {c.name: getattr(t, c.name) for c in t.__table__.columns}

def _staff_dict(s):
    return {c.name: getattr(s, c.name) for c in s.__table__.columns}

def _owner_dict(o):
    return {c.name: getattr(o, c.name) for c in o.__table__.columns}

def _v_dict(v):
    return {c.name: getattr(v, c.name) for c in v.__table__.columns}

def _org_dict(o):
    return {c.name: getattr(o, c.name) for c in o.__table__.columns}
